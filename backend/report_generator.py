import io
import base64
from datetime import datetime
from typing import Dict, Any, List
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from translations import get_translation


def generate_pdf_report(input_data: Dict[str, Any], predictions: Dict[str, Any], 
                       soil_health: Dict[str, Any], weather_data: Dict[str, Any] = None, 
                       language: str = 'en') -> bytes:
    """
    Generate a PDF report with prediction results and soil health analysis.
    
    Args:
        input_data: User input parameters
        predictions: ML model predictions
        soil_health: Soil health analysis results
        weather_data: Optional weather data
        language: Language code for translations (default: 'en')
        
    Returns:
        PDF content as bytes
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, 
                           topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center alignment
        textColor=colors.darkgreen
    )
    title = Paragraph(get_translation('report_title', language), title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Report date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=2,  # Right alignment
        textColor=colors.grey
    )
    date_text = f"{get_translation('generated_on', language)} {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    elements.append(Paragraph(date_text, date_style))
    elements.append(Spacer(1, 20))
    
    # Input Parameters Section
    elements.append(Paragraph(get_translation('input_parameters', language), styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    input_table_data = [
        ['Parameter', 'Value'],
        ['Temperature (°C)', str(input_data.get('Temparature', 'N/A'))],
        ['Humidity (%)', str(input_data.get('Humidity', 'N/A'))],
        ['Moisture (%)', str(input_data.get('Moisture', 'N/A'))],
        ['Soil Type', str(input_data.get('Soil_Type', 'N/A'))],
        ['Crop Type', str(input_data.get('Crop_Type', 'N/A'))],
        ['Nitrogen (N)', str(input_data.get('Nitrogen', 'N/A'))],
        ['Potassium (K)', str(input_data.get('Potassium', 'N/A'))],
        ['Phosphorus (P)', str(input_data.get('Phosphorous', 'N/A'))]
    ]
    
    input_table = Table(input_table_data, colWidths=[2*inch, 2*inch])
    input_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(input_table)
    elements.append(Spacer(1, 20))
    
    # Weather Data Section (if available)
    if weather_data:
        elements.append(Paragraph(get_translation('weather_information', language), styles['Heading2']))
        elements.append(Spacer(1, 12))
        
        weather_text = f"""
        <b>Location:</b> {weather_data.get('city', 'N/A')}, {weather_data.get('country', 'N/A')}<br/>
        <b>Temperature:</b> {weather_data.get('temperature', 'N/A')}°C<br/>
        <b>Humidity:</b> {weather_data.get('humidity', 'N/A')}%<br/>
        <b>Description:</b> {weather_data.get('description', 'N/A')}<br/>
        <b>Rainfall (1h):</b> {weather_data.get('rainfall', 0)} mm
        """
        elements.append(Paragraph(weather_text, styles['Normal']))
        elements.append(Spacer(1, 20))
    
    # Predictions Section
    elements.append(Paragraph(get_translation('fertilizer_predictions', language), styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    pred_table_data = [['Model', 'Predicted Fertilizer', get_translation('confidence', language)]]
    for model_name, result in predictions.items():
        confidence = max(result.get('probabilities', {}).values()) if result.get('probabilities') else 0
        pred_table_data.append([
            model_name,
            result.get('prediction', 'N/A'),
            f"{confidence:.2%}"
        ])
    
    pred_table = Table(pred_table_data, colWidths=[2*inch, 2*inch, 1.5*inch])
    pred_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(pred_table)
    elements.append(Spacer(1, 20))
    
    # Soil Health Section
    elements.append(Paragraph(get_translation('soil_health_analysis', language), styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    health_score = soil_health.get('health_score', 0)
    status = soil_health.get('overall_status', 'Unknown')
    
    health_text = f"""
    <b>{get_translation('overall_health_score', language)}:</b> {health_score}/100<br/>
    <b>{get_translation('status', language)}:</b> {status}<br/><br/>
    <b>{get_translation('key_insights', language)}:</b><br/>
    """
    
    for insight in soil_health.get('insights', []):
        health_text += f"• {insight}<br/>"
    
    health_text += f"<br/><b>{get_translation('recommendations', language)}:</b><br/>"
    for rec in soil_health.get('recommendations', []):
        health_text += f"• {rec}<br/>"
    
    elements.append(Paragraph(health_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Footer
    footer_text = get_translation('report_footer', language)
    elements.append(Paragraph(footer_text, date_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(input_data: Dict[str, Any], predictions: Dict[str, Any], 
                         soil_health: Dict[str, Any], weather_data: Dict[str, Any] = None, 
                         language: str = 'en') -> bytes:
    """
    Generate an Excel report with prediction results and soil health analysis.
    
    Args:
        input_data: User input parameters
        predictions: ML model predictions
        soil_health: Soil health analysis results
        weather_data: Optional weather data
        language: Language code for translations (default: 'en')
        
    Returns:
        Excel content as bytes
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Input Parameters Sheet
    ws_input = wb.create_sheet(get_translation('input_parameters', language))
    ws_input.append(["Parameter", "Value"])
    
    input_params = [
        ("Temperature (°C)", input_data.get('Temparature', 'N/A')),
        ("Humidity (%)", input_data.get('Humidity', 'N/A')),
        ("Moisture (%)", input_data.get('Moisture', 'N/A')),
        ("Soil Type", input_data.get('Soil_Type', 'N/A')),
        ("Crop Type", input_data.get('Crop_Type', 'N/A')),
        ("Nitrogen (N)", input_data.get('Nitrogen', 'N/A')),
        ("Potassium (K)", input_data.get('Potassium', 'N/A')),
        ("Phosphorus (P)", input_data.get('Phosphorous', 'N/A'))
    ]
    
    for param, value in input_params:
        ws_input.append([param, value])
    
    # Style input sheet
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    
    for cell in ws_input[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Predictions Sheet
    ws_pred = wb.create_sheet(get_translation('fertilizer_predictions', language))
    ws_pred.append(["Model", "Predicted Fertilizer", get_translation('confidence', language)])
    
    for model_name, result in predictions.items():
        confidence = max(result.get('probabilities', {}).values()) if result.get('probabilities') else 0
        ws_pred.append([
            model_name,
            result.get('prediction', 'N/A'),
            confidence
        ])
    
    # Style predictions sheet
    for cell in ws_pred[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add confidence chart
    chart = BarChart()
    chart.title = get_translation('model_confidence_comparison', language)
    chart.y_axis.title = get_translation('confidence', language)
    chart.x_axis.title = get_translation('models', language)
    
    data = Reference(ws_pred, min_col=3, min_row=2, max_row=len(predictions)+1)
    cats = Reference(ws_pred, min_col=1, min_row=2, max_row=len(predictions)+1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    
    ws_pred.add_chart(chart, "E2")
    
    # Soil Health Sheet
    ws_health = wb.create_sheet(get_translation('soil_health_analysis', language))
    ws_health.append([get_translation('soil_health_analysis', language)])
    ws_health.append([])
    ws_health.append([get_translation('overall_health_score', language), soil_health.get('health_score', 0)])
    ws_health.append([get_translation('status', language), soil_health.get('overall_status', 'Unknown')])
    ws_health.append([])
    ws_health.append([get_translation('key_insights', language)])
    
    for insight in soil_health.get('insights', []):
        ws_health.append([insight])
    
    ws_health.append([])
    ws_health.append([get_translation('recommendations', language)])
    
    for rec in soil_health.get('recommendations', []):
        ws_health.append([rec])
    
    # Style health sheet
    ws_health['A1'].font = Font(bold=True, size=14)
    ws_health['A7'].font = Font(bold=True)
    ws_health['A9'].font = Font(bold=True)
    
    # Weather Sheet (if available)
    if weather_data:
        ws_weather = wb.create_sheet(get_translation('weather_information', language))
        ws_weather.append([get_translation('weather_information', language)])
        ws_weather.append([])
        ws_weather.append(["Location", f"{weather_data.get('city', 'N/A')}, {weather_data.get('country', 'N/A')}"])
        ws_weather.append(["Temperature (°C)", weather_data.get('temperature', 'N/A')])
        ws_weather.append(["Humidity (%)", weather_data.get('humidity', 'N/A')])
        ws_weather.append(["Description", weather_data.get('description', 'N/A')])
        ws_weather.append(["Rainfall (1h mm)", weather_data.get('rainfall', 0)])
        
        ws_weather['A1'].font = Font(bold=True, size=14)
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append([get_translation('report_title', language)])
    ws_summary.append([f"{get_translation('generated_on', language)} {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"])
    ws_summary.append([])
    ws_summary.append([get_translation('report_contents', language)])
    ws_summary.append([f"1. {get_translation('input_parameters', language)} - {get_translation('input_parameters_desc', language)}"])
    ws_summary.append([f"2. {get_translation('fertilizer_predictions', language)} - {get_translation('predictions_desc', language)}"])
    ws_summary.append([f"3. {get_translation('soil_health_analysis', language)} - {get_translation('soil_health_desc', language)}"])
    if weather_data:
        ws_summary.append([f"4. {get_translation('weather_information', language)} - {get_translation('weather_data_desc', language)}"])
    ws_summary.append([f"5. Summary - {get_translation('summary_desc', language)}"])
    
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'].font = Font(italic=True)
    ws_summary['A4'].font = Font(bold=True)
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
